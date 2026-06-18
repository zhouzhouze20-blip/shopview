#!/usr/bin/env python3
"""
Sync Enterprise WeChat contacts into ShopView users and wecom identities.

Default mode is dry-run. Add --apply to write changes.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

sys.path.append(str(Path(__file__).resolve().parent))

from models.database import SessionLocal
from models.models import (
    DataPolicy,
    DataPolicyItem,
    Department,
    LoginLog,
    OperationLog,
    Post,
    Role,
    SystemLog,
    Store,
    User,
    UserDepartmentPost,
    UserIdentity,
    UserRole,
    WeComRoleScopeRule,
)
from routers.auth import DEFAULT_ADMIN_PASSWORD, _hash_password
WECOM_API_BASE = "https://qyapi.weixin.qq.com/cgi-bin"
DEFAULT_TIMEOUT_SECONDS = 12
WECOM_SOURCE_TYPE = "WECOM"
WECOM_SOURCE_SYSTEM = "wecom"
BUSINESS_SCOPE_RESOURCE = "business_scope"
BUSINESS_SCOPE_ACTION = "view"
DEFAULT_SYNC_ROLE_CODE = "dept_manager"
DEFAULT_ALLOWED_DEPARTMENT_KEYWORDS = ["百货条线", "集团总裁办"]
DEFAULT_ROLE_SCOPE_RULES = [
    {
        "position_keywords": ["店长", "总经理", "门店管理员"],
        "role_codes": ["store_admin"],
        "scope_mode": "CUSTOM",
        "scope_dimensions": {"store": ["$store_id"]},
    },
    {
        "position_keywords": ["部门主管", "部门经理", "主管", "经理", "副经理"],
        "role_codes": ["dept_manager"],
        "scope_mode": "CUSTOM",
        "scope_dimensions": {"department": ["$department"]},
    },
    {
        "position_keywords": ["柜组负责人", "柜组主管", "柜组长", "柜长", "组长"],
        "role_codes": ["group_manager"],
        "scope_mode": "CUSTOM",
        "scope_dimensions": {"department": ["$department"]},
    },
    {
        "department_keywords": ["财务"],
        "position_keywords": ["财务"],
        "match_any": True,
        "role_codes": ["finance"],
        "scope_mode": "CUSTOM",
        "scope_dimensions": {"store": ["$store_id"]},
    },
]


class WeComSyncError(RuntimeError):
    pass


def _log(message: str) -> None:
    print(message, flush=True)


@dataclass(frozen=True)
class WeComMember:
    userid: str
    name: str
    mobile: str
    email: str
    status: int
    department: str
    position: str
    raw: dict[str, Any]


def _request_json(url: str, params: dict[str, Any]) -> dict[str, Any]:
    try:
        response = requests.get(url, params=params, timeout=DEFAULT_TIMEOUT_SECONDS)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise WeComSyncError(f"企业微信接口请求失败: {exc}") from exc

    payload = response.json()
    errcode = int(payload.get("errcode", 0) or 0)
    if errcode != 0:
        errmsg = payload.get("errmsg") or "unknown error"
        raise WeComSyncError(f"企业微信接口返回错误: {errcode} {errmsg}")
    return payload


def _get_env(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def get_access_token(corp_id: str, secret: str) -> str:
    payload = _request_json(
        f"{WECOM_API_BASE}/gettoken",
        {"corpid": corp_id, "corpsecret": secret},
    )
    access_token = str(payload.get("access_token") or "")
    if not access_token:
        raise WeComSyncError("企业微信接口未返回 access_token")
    return access_token


def list_departments(access_token: str) -> tuple[list[int], dict[int, str]]:
    _log("fetching Enterprise WeChat departments...")
    payload = _request_json(f"{WECOM_API_BASE}/department/list", {"access_token": access_token})
    department_names: dict[int, str] = {}
    parent_ids: dict[int, int] = {}
    ids: list[int] = []
    for item in payload.get("department", []):
        if item.get("id") is None:
            continue
        department_id = int(item["id"])
        ids.append(department_id)
        department_names[department_id] = str(item.get("name") or item.get("name_en") or department_id).strip()
        if item.get("parentid") is not None:
            parent_ids[department_id] = int(item.get("parentid") or 0)

    def department_path(department_id: int) -> str:
        path: list[str] = []
        seen: set[int] = set()
        current_id = department_id
        while current_id and current_id not in seen:
            seen.add(current_id)
            name = department_names.get(current_id, "").strip()
            if name:
                path.append(name)
            current_id = parent_ids.get(current_id, 0)
        path.reverse()
        return "/".join(path)

    department_paths = {
        department_id: department_path(department_id) or department_names.get(department_id, str(department_id))
        for department_id in ids
    }
    _log(f"departments fetched: {len(ids or [1])}")
    return ids or [1], department_paths


def list_department_ids(access_token: str) -> list[int]:
    ids, _department_names = list_departments(access_token)
    return ids


def list_members(access_token: str, department_ids: list[int], department_names: dict[int, str] | None = None) -> list[WeComMember]:
    members_by_userid: dict[str, WeComMember] = {}
    department_names = department_names or {}
    for index, department_id in enumerate(department_ids, start=1):
        _log(f"fetching members for department {department_id} ({index}/{len(department_ids)})...")
        payload = _request_json(
            f"{WECOM_API_BASE}/user/list",
            {
                "access_token": access_token,
                "department_id": department_id,
                "fetch_child": 0,
            },
        )
        for row in payload.get("userlist", []):
            userid = str(row.get("userid") or "").strip()
            if not userid:
                continue
            row_department_names = [
                department_names.get(int(row_department_id), str(row_department_id))
                for row_department_id in row.get("department", []) or []
            ]
            members_by_userid[userid] = WeComMember(
                userid=userid,
                name=str(row.get("name") or "").strip(),
                mobile=str(row.get("mobile") or "").strip(),
                email=str(row.get("email") or "").strip(),
                status=int(row.get("status", 1) or 1),
                department=";".join(name for name in row_department_names if name),
                position=str(row.get("position") or "").strip(),
                raw=row,
            )
        _log(f"members accumulated: {len(members_by_userid)}")
    return sorted(members_by_userid.values(), key=lambda item: item.userid)


def _norm_header(value: object) -> str:
    return str(value or "").strip().replace(" ", "").replace("\u3000", "").lower()


def _pick(row: dict[str, str], names: list[str]) -> str:
    normalized = {_norm_header(key): value for key, value in row.items()}
    for name in names:
        value = normalized.get(_norm_header(name), "")
        if value:
            return value.strip()
    return ""


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def _load_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WeComSyncError("缺少 openpyxl，无法读取 xlsx 文件") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result: list[dict[str, str]] = []
    for values in rows[1:]:
        result.append({
            headers[index]: str(value or "").strip()
            for index, value in enumerate(values)
            if index < len(headers)
        })
    return result


def load_members_from_file(path: str | Path) -> list[WeComMember]:
    file_path = Path(path)
    if not file_path.exists():
        raise WeComSyncError(f"通讯录文件不存在: {file_path}")

    _log(f"loading contacts from file: {file_path}")
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        rows = _load_csv_rows(file_path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _load_xlsx_rows(file_path)
    else:
        raise WeComSyncError("仅支持 .csv/.xlsx 通讯录文件")

    members: list[WeComMember] = []
    for row in rows:
        userid = _pick(row, ["账号", "帐号", "userid", "user_id", "成员UserID", "成员账号", "工号"])
        name = _pick(row, ["姓名", "名称", "成员名称", "name"])
        if not userid:
            continue
        members.append(
            WeComMember(
                userid=userid,
                name=name or userid,
                mobile=_pick(row, ["手机", "手机号", "mobile"]),
                email=_pick(row, ["邮箱", "email"]),
                status=1,
                department=_pick(row, ["部门", "所在部门", "部门名称", "department"]),
                position=_pick(row, ["岗位", "职位", "职务", "position", "职务信息"]),
                raw=row,
            )
        )
    if not members:
        raise WeComSyncError("通讯录文件未识别到成员，请检查是否包含账号/UserID列")
    _log(f"members loaded from file: {len(members)}")
    return members


def _is_active_status(status: int) -> bool:
    return status == 1


def _find_user_for_member(db, member: WeComMember) -> User | None:
    user = db.query(User).filter(User.username == member.userid).first()
    if user:
        return user
    return db.query(User).filter(User.employee_no == member.userid).first()


def _upsert_user(db, member: WeComMember, default_password: str) -> tuple[User, bool]:
    user = _find_user_for_member(db, member)
    active = _is_active_status(member.status)
    if user:
        user.real_name = member.name or user.real_name
        user.phone = member.mobile or user.phone
        user.email = member.email or user.email
        user.employee_no = member.userid
        user.status = "ACTIVE" if active else "DISABLED"
        user.is_active = active
        user.updated_at = datetime.now()
        return user, False

    user = User(
        username=member.userid,
        password_hash=_hash_password(default_password),
        real_name=member.name or member.userid,
        phone=member.mobile or None,
        email=member.email or None,
        employee_no=member.userid,
        role="user",
        status="ACTIVE" if active else "DISABLED",
        is_active=active,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(user)
    db.flush()
    return user, True


def _upsert_wecom_identity(db, user: User, corp_id: str, member: WeComMember) -> bool:
    identity = (
        db.query(UserIdentity)
        .filter(
            UserIdentity.identity_type == "wecom",
            UserIdentity.corp_id == corp_id,
            UserIdentity.wecom_user_id == member.userid,
        )
        .first()
    )
    created = False
    if not identity:
        identity = UserIdentity(
            user_id=user.user_id,
            identity_type="wecom",
            identifier=f"{corp_id}:{member.userid}",
            corp_id=corp_id,
            wecom_user_id=member.userid,
            is_primary=False,
            created_at=datetime.now(),
        )
        db.add(identity)
        created = True

    identity.user_id = user.user_id
    identity.identifier = f"{corp_id}:{member.userid}"
    identity.updated_at = datetime.now()
    return created


def _dept_code(department_path: str) -> str:
    value = department_path.strip().strip("/")
    return f"wecom:{value}"[:50] if value else ""


def _dept_name(department_path: str) -> str:
    value = department_path.strip().strip("/")
    if not value:
        return ""
    parts = [part.strip() for part in value.replace("\\", "/").split("/") if part.strip()]
    return parts[-1] if parts else value


def _department_leaf_names(department_path: str) -> list[str]:
    chunks = [
        chunk.strip()
        for chunk in department_path.replace("；", ";").replace("，", ";").replace(",", ";").split(";")
        if chunk.strip()
    ]
    names: list[str] = []
    for chunk in chunks:
        name = _dept_name(chunk)
        if name and name not in names:
            names.append(name)
    return names


def _department_scope_values(department_path: str) -> list[str]:
    values: list[str] = []
    paths = [
        path.strip()
        for path in department_path.replace("；", ";").replace("，", ";").replace(",", ";").split(";")
        if path.strip()
    ]
    for path in paths:
        for part in path.replace("\\", "/").split("/"):
            value = part.strip()
            if value and value not in values:
                values.append(value)
    return values


def _norm_text(value: object) -> str:
    return str(value or "").strip().lower()


def _parse_keyword_list(raw: str) -> list[str]:
    return [
        item.strip()
        for item in raw.replace("；", ",").replace(";", ",").split(",")
        if item.strip()
    ]


def _allowed_department_keywords(argument_value: str = "") -> list[str]:
    raw = argument_value.strip() or _get_env("WECOM_ALLOWED_DEPARTMENT_KEYWORDS")
    if not raw:
        return DEFAULT_ALLOWED_DEPARTMENT_KEYWORDS[:]
    if raw.strip() in {"*", "ALL", "all"}:
        return []
    return _parse_keyword_list(raw)


def _filter_members_by_departments(
    members: list[WeComMember],
    allowed_department_keywords: list[str],
) -> tuple[list[WeComMember], int]:
    if not allowed_department_keywords:
        return members, 0
    filtered = [
        member
        for member in members
        if any(keyword in member.department for keyword in allowed_department_keywords)
    ]
    return filtered, len(members) - len(filtered)


def _rule_record_to_dict(rule: WeComRoleScopeRule) -> dict[str, Any]:
    payload = {
        "userids": rule.wecom_userids or [],
        "names": rule.name_keywords or [],
        "department_keywords": rule.department_keywords or [],
        "position_keywords": rule.position_keywords or [],
        "role_codes": rule.role_codes or [],
        "scope_mode": rule.scope_mode,
        "scope_dimensions": rule.scope_dimensions or {},
    }
    if rule.match_mode == "ANY":
        payload["match_any"] = True
    return payload


def _load_role_scope_rules(path: str = "", db=None) -> list[dict[str, Any]]:
    if db is not None:
        rules = (
            db.query(WeComRoleScopeRule)
            .filter(WeComRoleScopeRule.is_active == True)
            .order_by(WeComRoleScopeRule.priority.asc(), WeComRoleScopeRule.id.asc())
            .all()
        )
        if rules:
            return [_rule_record_to_dict(rule) for rule in rules]

    raw = Path(path).read_text(encoding="utf-8") if path else _get_env("WECOM_ROLE_SCOPE_RULES")
    if not raw:
        return []
    payload = json.loads(raw)
    if not isinstance(payload, list):
        raise WeComSyncError("企业微信角色范围规则必须是 JSON 数组")
    return [item for item in payload if isinstance(item, dict)]


def _rule_matches(member: WeComMember, rule: dict[str, Any]) -> bool:
    checks: list[bool] = []
    userids = [_norm_text(item) for item in rule.get("userids", []) if _norm_text(item)]
    if userids:
        checks.append(_norm_text(member.userid) in userids)
    names = [_norm_text(item) for item in rule.get("names", []) if _norm_text(item)]
    if names:
        checks.append(_norm_text(member.name) in names)
    department_keywords = [_norm_text(item) for item in rule.get("department_keywords", []) if _norm_text(item)]
    if department_keywords:
        department_text = _norm_text(member.department)
        checks.append(any(keyword in department_text for keyword in department_keywords))
    position_keywords = [_norm_text(item) for item in rule.get("position_keywords", []) if _norm_text(item)]
    if position_keywords:
        position_text = _norm_text(member.position)
        checks.append(any(keyword in position_text for keyword in position_keywords))
    if not checks:
        return False
    return any(checks) if rule.get("match_any") else all(checks)


def _expand_scope_value(template: object, *, store_id: int, member: WeComMember) -> list[str]:
    value = str(template or "").strip()
    if not value:
        return []
    if value == "$store_id":
        return [str(store_id)]
    if value == "$department":
        return _department_leaf_names(member.department)
    if value == "$department_path":
        return [member.department.strip()] if member.department.strip() else []
    if value == "$userid":
        return [member.userid]
    return [value]


def _resolve_role_scope(
    member: WeComMember,
    *,
    store_id: int,
    rules: list[dict[str, Any]],
    use_default_rules: bool,
) -> tuple[list[str], str, dict[str, list[str]]]:
    effective_rules = rules or (DEFAULT_ROLE_SCOPE_RULES if use_default_rules else [])
    role_codes: list[str] = []
    scope_mode = "NONE"
    scope_dimensions: dict[str, list[str]] = {}

    for rule in effective_rules:
        if not _rule_matches(member, rule):
            continue
        for role_code in rule.get("role_codes", []):
            role_code = str(role_code or "").strip()
            if role_code and role_code not in role_codes:
                role_codes.append(role_code)
        rule_scope_mode = str(rule.get("scope_mode") or "CUSTOM").strip().upper()
        if rule_scope_mode == "ALL":
            scope_mode = "ALL"
            scope_dimensions.clear()
            continue
        if scope_mode != "ALL" and rule_scope_mode == "CUSTOM":
            scope_mode = "CUSTOM"
            for dimension_type, values in (rule.get("scope_dimensions") or {}).items():
                dimension_type = str(dimension_type or "").strip()
                if not dimension_type:
                    continue
                for template in values or []:
                    for value in _expand_scope_value(template, store_id=store_id, member=member):
                        scope_dimensions.setdefault(dimension_type, [])
                        if value not in scope_dimensions[dimension_type]:
                            scope_dimensions[dimension_type].append(value)

    department_scope_values = _department_scope_values(member.department)

    if not role_codes and use_default_rules:
        role_codes.append(DEFAULT_SYNC_ROLE_CODE)

    if department_scope_values and scope_mode != "ALL":
        scope_mode = "CUSTOM"
        scope_dimensions = {"department": department_scope_values}

    return role_codes, scope_mode, scope_dimensions


def _ensure_post(db, position: str) -> Post | None:
    position = position.strip()
    if not position:
        return None
    post_code = f"wecom:{position}"[:50]
    post = db.query(Post).filter(Post.post_code == post_code).first()
    if post:
        post.post_name = position
        post.is_active = True
        post.updated_at = datetime.now()
        return post
    post = Post(
        post_code=post_code,
        post_name=position,
        level=0,
        is_active=True,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(post)
    db.flush()
    return post


def _sync_user_department(db, user: User, member: WeComMember) -> bool:
    if not member.department:
        return False
    store = db.query(Store).order_by(Store.store_id.asc()).first()
    if not store:
        raise WeComSyncError("无法同步部门：stores 表为空")

    dept_code = _dept_code(member.department)
    department = (
        db.query(Department)
        .filter(Department.store_id == store.store_id, Department.dept_code == dept_code)
        .first()
    )
    if not department:
        department = Department(
            store_id=store.store_id,
            dept_code=dept_code,
            dept_name=_dept_name(member.department),
            is_active=True,
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
        db.add(department)
        db.flush()
    else:
        department.dept_name = _dept_name(member.department) or department.dept_name
        department.is_active = True
        department.updated_at = datetime.now()

    post = _ensure_post(db, member.position)
    existing = (
        db.query(UserDepartmentPost)
        .filter(
            UserDepartmentPost.user_id == user.user_id,
            UserDepartmentPost.department_id == department.id,
        )
        .first()
    )
    if existing:
        existing.post_id = post.id if post else existing.post_id
        existing.is_primary = True
        existing.is_active = True
        existing.updated_at = datetime.now()
        return False

    db.add(
        UserDepartmentPost(
            user_id=user.user_id,
            store_id=store.store_id,
            department_id=department.id,
            post_id=post.id if post else None,
            is_primary=True,
            is_active=True,
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
    )
    return True


def _ensure_viewer_role(db, user_id: int) -> bool:
    role = db.query(Role).filter(Role.role_code == DEFAULT_SYNC_ROLE_CODE, Role.is_active == True).first()
    if not role:
        raise WeComSyncError(f"默认同步角色不存在或未启用: {DEFAULT_SYNC_ROLE_CODE}")
    exists = db.query(UserRole).filter(UserRole.user_id == user_id, UserRole.role_id == role.id).first()
    if exists:
        return False
    db.add(UserRole(user_id=user_id, role_id=role.id, created_at=datetime.now()))
    return True


def _ensure_user_role(db, user_id: int, role_code: str) -> bool:
    if role_code == "contract_viewer":
        role_code = DEFAULT_SYNC_ROLE_CODE
    role = db.query(Role).filter(Role.role_code == role_code, Role.is_active == True).first()
    if not role:
        raise WeComSyncError(f"自动分配角色失败：角色不存在或未启用 {role_code}")
    exists = db.query(UserRole).filter(UserRole.user_id == user_id, UserRole.role_id == role.id).first()
    if exists:
        return False
    db.add(UserRole(user_id=user_id, role_id=role.id, created_at=datetime.now()))
    return True


def _refresh_wecom_business_scope(
    db,
    user_id: int,
    member: WeComMember,
    scope_mode: str,
    scope_dimensions: dict[str, list[str]],
) -> bool:
    existing_policy_ids = [
        row[0]
        for row in db.query(DataPolicy.id)
        .filter(
            DataPolicy.subject_type == "USER",
            DataPolicy.subject_id == user_id,
            DataPolicy.resource_code == BUSINESS_SCOPE_RESOURCE,
            DataPolicy.action_code == BUSINESS_SCOPE_ACTION,
            DataPolicy.source_type == WECOM_SOURCE_TYPE,
            DataPolicy.source_system == WECOM_SOURCE_SYSTEM,
        )
        .all()
    ]
    if existing_policy_ids:
        db.query(DataPolicyItem).filter(DataPolicyItem.policy_id.in_(existing_policy_ids)).delete(synchronize_session=False)
        db.query(DataPolicy).filter(DataPolicy.id.in_(existing_policy_ids)).delete(synchronize_session=False)

    scope_mode = scope_mode.upper()
    if scope_mode not in {"ALL", "CUSTOM"}:
        return bool(existing_policy_ids)
    if scope_mode == "CUSTOM" and not any(scope_dimensions.values()):
        return bool(existing_policy_ids)

    policy = DataPolicy(
        subject_type="USER",
        subject_id=user_id,
        resource_code=BUSINESS_SCOPE_RESOURCE,
        action_code=BUSINESS_SCOPE_ACTION,
        scope_mode=scope_mode,
        effect="ALLOW",
        priority=60,
        is_active=True,
        source_type=WECOM_SOURCE_TYPE,
        source_system=WECOM_SOURCE_SYSTEM,
        external_scope_id=f"{member.userid}:business_scope",
        external_scope_name=f"{member.name or member.userid} 企业微信数据范围",
        synced_at=datetime.now(),
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(policy)
    db.flush()

    if scope_mode == "CUSTOM":
        for dimension_type, values in scope_dimensions.items():
            for value in values:
                db.add(
                    DataPolicyItem(
                        policy_id=policy.id,
                        dimension_type=dimension_type,
                        dimension_value=str(value),
                        include_children=False,
                        created_at=datetime.now(),
                    )
                )
    return True


def _sync_role_scope_for_member(
    db,
    user: User,
    member: WeComMember,
    *,
    rules: list[dict[str, Any]],
    use_default_rules: bool,
) -> Counter:
    stats = Counter()
    store = db.query(Store).order_by(Store.store_id.asc()).first()
    if not store:
        raise WeComSyncError("无法自动分配角色和数据范围：stores 表为空")

    role_codes, scope_mode, scope_dimensions = _resolve_role_scope(
        member,
        store_id=store.store_id,
        rules=rules,
        use_default_rules=use_default_rules,
    )
    if not role_codes and scope_mode == "NONE":
        return stats

    for role_code in role_codes:
        if _ensure_user_role(db, user.user_id, role_code):
            stats["auto_roles_added"] += 1

    if _refresh_wecom_business_scope(db, user.user_id, member, scope_mode, scope_dimensions):
        stats["wecom_business_scopes_refreshed"] += 1
    return stats


def _reset_non_admin_users(db) -> Counter:
    _log("resetting non-admin users...")
    stats = Counter()
    admin = db.query(User).filter(User.username == "admin").first()
    admin_user_id = admin.user_id if admin else None

    query = db.query(User.user_id)
    if admin_user_id is not None:
        query = query.filter(User.user_id != admin_user_id)
    user_ids = [row[0] for row in query.all()]
    if not user_ids:
        _log("no non-admin users to reset")
        return stats
    _log(f"non-admin users to delete: {len(user_ids)}")

    policy_ids = [
        row[0]
        for row in db.query(DataPolicy.id)
        .filter(DataPolicy.subject_type == "USER", DataPolicy.subject_id.in_(user_ids))
        .all()
    ]
    if policy_ids:
        stats["data_policy_items_deleted"] = db.query(DataPolicyItem).filter(
            DataPolicyItem.policy_id.in_(policy_ids)
        ).delete(synchronize_session=False)
        stats["data_policies_deleted"] = db.query(DataPolicy).filter(
            DataPolicy.id.in_(policy_ids)
        ).delete(synchronize_session=False)

    stats["department_managers_cleared"] = db.query(Department).filter(
        Department.manager_user_id.in_(user_ids)
    ).update({Department.manager_user_id: None}, synchronize_session=False)
    stats["operation_logs_unlinked"] = db.query(OperationLog).filter(
        OperationLog.user_id.in_(user_ids)
    ).update({OperationLog.user_id: None}, synchronize_session=False)
    stats["system_logs_unlinked"] = db.query(SystemLog).filter(
        SystemLog.user_id.in_(user_ids)
    ).update({SystemLog.user_id: None}, synchronize_session=False)
    stats["login_logs_deleted"] = db.query(LoginLog).filter(
        LoginLog.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["user_department_posts_deleted"] = db.query(UserDepartmentPost).filter(
        UserDepartmentPost.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["user_roles_deleted"] = db.query(UserRole).filter(
        UserRole.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["user_identities_deleted"] = db.query(UserIdentity).filter(
        UserIdentity.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["users_deleted"] = db.query(User).filter(
        User.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    return stats


def sync_wecom_contacts(
    *,
    apply: bool,
    from_file: str,
    reset_users: bool,
    reset_wecom_identities: bool,
    disable_missing_wecom_users: bool,
    grant_viewer_role: bool,
    auto_role_scope: bool,
    role_scope_rules: str,
    allowed_departments: str,
    default_password: str,
) -> int:
    corp_id = _get_env("WECOM_CORP_ID")
    if not corp_id:
        raise WeComSyncError("缺少 WECOM_CORP_ID")

    if from_file:
        department_ids: list[int] = []
        members = load_members_from_file(from_file)
    else:
        secret = _get_env("WECOM_CONTACT_SECRET") or _get_env("WECOM_APP_SECRET")
        if not secret:
            raise WeComSyncError("缺少 WECOM_CONTACT_SECRET 或 WECOM_APP_SECRET")
        _log("fetching Enterprise WeChat access token...")
        access_token = get_access_token(corp_id, secret)
        _log("access token fetched")
        department_ids, department_names = list_departments(access_token)
        members = list_members(access_token, department_ids, department_names)
    members_seen_before_filter = len(members)
    allowed_department_keywords = _allowed_department_keywords(allowed_departments)
    members, filtered_out_count = _filter_members_by_departments(members, allowed_department_keywords)
    active_userids = {member.userid for member in members if _is_active_status(member.status)}
    db = SessionLocal()
    stats = Counter()
    try:
        rules = _load_role_scope_rules(role_scope_rules, db=db)
        db_rule_count = db.query(WeComRoleScopeRule).filter(WeComRoleScopeRule.is_active == True).count()
        stats["departments_seen"] = len(department_ids)
        stats["members_seen"] = members_seen_before_filter
        stats["members_filtered_out"] = filtered_out_count
        stats["members_active"] = len(active_userids)
        stats["role_scope_rules"] = len(rules)

        if apply and reset_users:
            stats.update(_reset_non_admin_users(db))

        if apply and reset_wecom_identities:
            deleted = db.query(UserIdentity).filter(UserIdentity.identity_type == "wecom").delete(synchronize_session=False)
            stats["wecom_identities_deleted"] = deleted

        _log(f"syncing members into database: {len(members)}")
        for index, member in enumerate(members, start=1):
            if index == 1 or index % 20 == 0 or index == len(members):
                _log(f"sync progress: {index}/{len(members)}")
            if not _is_active_status(member.status):
                stats["members_inactive"] += 1

            if not apply:
                user = _find_user_for_member(db, member)
                stats["users_matched" if user else "users_would_create"] += 1
                continue

            user, created = _upsert_user(db, member, default_password)
            stats["users_created" if created else "users_updated"] += 1
            if _upsert_wecom_identity(db, user, corp_id, member):
                stats["wecom_identities_created"] += 1
            else:
                stats["wecom_identities_updated"] += 1
            if grant_viewer_role and _ensure_viewer_role(db, user.user_id):
                stats["viewer_roles_added"] += 1
            if _sync_user_department(db, user, member):
                stats["user_departments_created"] += 1
            if auto_role_scope or rules or member.department:
                stats.update(
                    _sync_role_scope_for_member(
                        db,
                        user,
                        member,
                        rules=rules,
                        use_default_rules=auto_role_scope and not rules,
                    )
                )

        if apply and disable_missing_wecom_users:
            identities = (
                db.query(UserIdentity)
                .filter(UserIdentity.identity_type == "wecom", UserIdentity.corp_id == corp_id)
                .all()
            )
            for identity in identities:
                if identity.wecom_user_id in active_userids:
                    continue
                user = db.query(User).filter(User.user_id == identity.user_id).first()
                if user and user.username != "admin":
                    user.status = "DISABLED"
                    user.is_active = False
                    user.updated_at = datetime.now()
                    stats["users_disabled_missing_wecom"] += 1

        if apply:
            db.commit()
        else:
            db.rollback()

        print("Enterprise WeChat contacts sync summary")
        print(f"mode: {'APPLY' if apply else 'DRY-RUN'}")
        print(f"corp_id: {corp_id}")
        print(f"from_file: {from_file or '-'}")
        print(f"reset_users: {reset_users}")
        print(f"reset_wecom_identities: {reset_wecom_identities}")
        print(f"disable_missing_wecom_users: {disable_missing_wecom_users}")
        print(f"grant_viewer_role: {grant_viewer_role}")
        print(f"auto_role_scope: {auto_role_scope}")
        print(f"allowed_departments: {', '.join(allowed_department_keywords) if allowed_department_keywords else 'ALL'}")
        print(f"role_scope_rules: {'DB:wecom_role_scope_rules' if db_rule_count else role_scope_rules or ('ENV:WECOM_ROLE_SCOPE_RULES' if rules else '-')}")
        print("stats: " + (", ".join(f"{key}={value}" for key, value in sorted(stats.items())) or "none"))
        if not apply:
            print("dry-run only; add --apply to write changes")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Enterprise WeChat contacts into ShopView users.")
    parser.add_argument("--from-file", default="", help="Import members from Enterprise WeChat exported .csv/.xlsx instead of API.")
    parser.add_argument("--apply", action="store_true", help="Write changes. Omit for dry-run.")
    parser.add_argument("--reset-users", action="store_true", help="Delete all non-admin users and related user auth/data policy rows before syncing.")
    parser.add_argument("--reset-wecom-identities", action="store_true", help="Delete existing wecom identities before syncing.")
    parser.add_argument("--disable-missing-wecom-users", action="store_true", help="Disable users bound to wecom ids no longer active.")
    parser.add_argument("--grant-viewer-role", action="store_true", help="Grant the contract viewer role to synced users.")
    parser.add_argument("--auto-role-scope", action="store_true", help="Automatically assign roles and WECOM business scope from department and position.")
    parser.add_argument("--role-scope-rules", default="", help="JSON file for Enterprise WeChat role/scope rules. If omitted, WECOM_ROLE_SCOPE_RULES env can be used.")
    parser.add_argument("--allowed-departments", default="", help="Comma-separated department keywords to sync. Defaults to WECOM_ALLOWED_DEPARTMENT_KEYWORDS or 百货条线,集团总裁办. Use * for all.")
    parser.add_argument("--default-password", default=DEFAULT_ADMIN_PASSWORD, help="Initial password for newly-created users.")
    args = parser.parse_args()

    try:
        return sync_wecom_contacts(
            apply=args.apply,
            from_file=args.from_file,
            reset_users=args.reset_users,
            reset_wecom_identities=args.reset_wecom_identities,
            disable_missing_wecom_users=args.disable_missing_wecom_users,
            grant_viewer_role=args.grant_viewer_role,
            auto_role_scope=args.auto_role_scope,
            role_scope_rules=args.role_scope_rules,
            allowed_departments=args.allowed_departments,
            default_password=args.default_password,
        )
    except WeComSyncError as exc:
        print(f"同步失败: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
