#!/usr/bin/env python3
"""
Sync Enterprise WeChat contacts into ShopView users and wecom identities.

Default mode is dry-run. Add --apply to write changes.
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

sys.path.append(str(Path(__file__).resolve().parent))

from models.database import SessionLocal
from models.models import (
    DataPolicy,
    DataPolicyItem,
    Department,
    LoginLog,
    OperationLog,
    SystemLog,
    Store,
    User,
    UserDepartmentPost,
    UserIdentity,
    UserRole,
)
from routers.auth import DEFAULT_ADMIN_PASSWORD, _hash_password
from routers.system_management import _ensure_contract_viewer_role


WECOM_API_BASE = "https://qyapi.weixin.qq.com/cgi-bin"
DEFAULT_TIMEOUT_SECONDS = 12


class WeComSyncError(RuntimeError):
    pass


def _log(message: str) -> None:
    print(message, flush=True)


@dataclass(frozen=True)
class WeComMember:
    userid: str
    name: str
    mobile: str
    email: str
    status: int
    department: str
    raw: dict[str, Any]


def _request_json(url: str, params: dict[str, Any]) -> dict[str, Any]:
    try:
        response = requests.get(url, params=params, timeout=DEFAULT_TIMEOUT_SECONDS)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise WeComSyncError(f"企业微信接口请求失败: {exc}") from exc

    payload = response.json()
    errcode = int(payload.get("errcode", 0) or 0)
    if errcode != 0:
        errmsg = payload.get("errmsg") or "unknown error"
        raise WeComSyncError(f"企业微信接口返回错误: {errcode} {errmsg}")
    return payload


def _get_env(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def get_access_token(corp_id: str, secret: str) -> str:
    payload = _request_json(
        f"{WECOM_API_BASE}/gettoken",
        {"corpid": corp_id, "corpsecret": secret},
    )
    access_token = str(payload.get("access_token") or "")
    if not access_token:
        raise WeComSyncError("企业微信接口未返回 access_token")
    return access_token


def list_department_ids(access_token: str) -> list[int]:
    _log("fetching Enterprise WeChat departments...")
    payload = _request_json(f"{WECOM_API_BASE}/department/list", {"access_token": access_token})
    ids = [int(item["id"]) for item in payload.get("department", []) if item.get("id") is not None]
    _log(f"departments fetched: {len(ids or [1])}")
    return ids or [1]


def list_members(access_token: str, department_ids: list[int]) -> list[WeComMember]:
    members_by_userid: dict[str, WeComMember] = {}
    for index, department_id in enumerate(department_ids, start=1):
        _log(f"fetching members for department {department_id} ({index}/{len(department_ids)})...")
        payload = _request_json(
            f"{WECOM_API_BASE}/user/list",
            {
                "access_token": access_token,
                "department_id": department_id,
                "fetch_child": 0,
            },
        )
        for row in payload.get("userlist", []):
            userid = str(row.get("userid") or "").strip()
            if not userid:
                continue
            members_by_userid[userid] = WeComMember(
                userid=userid,
                name=str(row.get("name") or "").strip(),
                mobile=str(row.get("mobile") or "").strip(),
                email=str(row.get("email") or "").strip(),
                status=int(row.get("status", 1) or 1),
                department="",
                raw=row,
            )
        _log(f"members accumulated: {len(members_by_userid)}")
    return sorted(members_by_userid.values(), key=lambda item: item.userid)


def _norm_header(value: object) -> str:
    return str(value or "").strip().replace(" ", "").replace("\u3000", "").lower()


def _pick(row: dict[str, str], names: list[str]) -> str:
    normalized = {_norm_header(key): value for key, value in row.items()}
    for name in names:
        value = normalized.get(_norm_header(name), "")
        if value:
            return value.strip()
    return ""


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def _load_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WeComSyncError("缺少 openpyxl，无法读取 xlsx 文件") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result: list[dict[str, str]] = []
    for values in rows[1:]:
        result.append({
            headers[index]: str(value or "").strip()
            for index, value in enumerate(values)
            if index < len(headers)
        })
    return result


def load_members_from_file(path: str | Path) -> list[WeComMember]:
    file_path = Path(path)
    if not file_path.exists():
        raise WeComSyncError(f"通讯录文件不存在: {file_path}")

    _log(f"loading contacts from file: {file_path}")
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        rows = _load_csv_rows(file_path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _load_xlsx_rows(file_path)
    else:
        raise WeComSyncError("仅支持 .csv/.xlsx 通讯录文件")

    members: list[WeComMember] = []
    for row in rows:
        userid = _pick(row, ["账号", "帐号", "userid", "user_id", "成员UserID", "成员账号", "工号"])
        name = _pick(row, ["姓名", "名称", "成员名称", "name"])
        if not userid:
            continue
        members.append(
            WeComMember(
                userid=userid,
                name=name or userid,
                mobile=_pick(row, ["手机", "手机号", "mobile"]),
                email=_pick(row, ["邮箱", "email"]),
                status=1,
                department=_pick(row, ["部门", "所在部门", "部门名称", "department"]),
                raw=row,
            )
        )
    if not members:
        raise WeComSyncError("通讯录文件未识别到成员，请检查是否包含账号/UserID列")
    _log(f"members loaded from file: {len(members)}")
    return members


def _is_active_status(status: int) -> bool:
    return status == 1


def _find_user_for_member(db, member: WeComMember) -> User | None:
    user = db.query(User).filter(User.username == member.userid).first()
    if user:
        return user
    return db.query(User).filter(User.employee_no == member.userid).first()


def _upsert_user(db, member: WeComMember, default_password: str) -> tuple[User, bool]:
    user = _find_user_for_member(db, member)
    active = _is_active_status(member.status)
    if user:
        user.real_name = member.name or user.real_name
        user.phone = member.mobile or user.phone
        user.email = member.email or user.email
        user.employee_no = member.userid
        user.status = "ACTIVE" if active else "DISABLED"
        user.is_active = active
        user.updated_at = datetime.now()
        return user, False

    user = User(
        username=member.userid,
        password_hash=_hash_password(default_password),
        real_name=member.name or member.userid,
        phone=member.mobile or None,
        email=member.email or None,
        employee_no=member.userid,
        role="user",
        status="ACTIVE" if active else "DISABLED",
        is_active=active,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(user)
    db.flush()
    return user, True


def _upsert_wecom_identity(db, user: User, corp_id: str, member: WeComMember) -> bool:
    identity = (
        db.query(UserIdentity)
        .filter(
            UserIdentity.identity_type == "wecom",
            UserIdentity.corp_id == corp_id,
            UserIdentity.wecom_user_id == member.userid,
        )
        .first()
    )
    created = False
    if not identity:
        identity = UserIdentity(
            user_id=user.user_id,
            identity_type="wecom",
            identifier=f"{corp_id}:{member.userid}",
            corp_id=corp_id,
            wecom_user_id=member.userid,
            is_primary=False,
            created_at=datetime.now(),
        )
        db.add(identity)
        created = True

    identity.user_id = user.user_id
    identity.identifier = f"{corp_id}:{member.userid}"
    identity.updated_at = datetime.now()
    return created


def _dept_code(department_path: str) -> str:
    value = department_path.strip().strip("/")
    return f"wecom:{value}"[:50] if value else ""


def _dept_name(department_path: str) -> str:
    value = department_path.strip().strip("/")
    if not value:
        return ""
    parts = [part.strip() for part in value.replace("\\", "/").split("/") if part.strip()]
    return parts[-1] if parts else value


def _sync_user_department(db, user: User, member: WeComMember) -> bool:
    if not member.department:
        return False
    store = db.query(Store).order_by(Store.store_id.asc()).first()
    if not store:
        raise WeComSyncError("无法同步部门：stores 表为空")

    dept_code = _dept_code(member.department)
    department = (
        db.query(Department)
        .filter(Department.store_id == store.store_id, Department.dept_code == dept_code)
        .first()
    )
    if not department:
        department = Department(
            store_id=store.store_id,
            dept_code=dept_code,
            dept_name=_dept_name(member.department),
            is_active=True,
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
        db.add(department)
        db.flush()
    else:
        department.dept_name = _dept_name(member.department) or department.dept_name
        department.is_active = True
        department.updated_at = datetime.now()

    existing = (
        db.query(UserDepartmentPost)
        .filter(
            UserDepartmentPost.user_id == user.user_id,
            UserDepartmentPost.department_id == department.id,
        )
        .first()
    )
    if existing:
        existing.is_primary = True
        existing.is_active = True
        existing.updated_at = datetime.now()
        return False

    db.add(
        UserDepartmentPost(
            user_id=user.user_id,
            store_id=store.store_id,
            department_id=department.id,
            is_primary=True,
            is_active=True,
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
    )
    return True


def _ensure_viewer_role(db, user_id: int) -> bool:
    role = _ensure_contract_viewer_role(db)
    exists = db.query(UserRole).filter(UserRole.user_id == user_id, UserRole.role_id == role.id).first()
    if exists:
        return False
    db.add(UserRole(user_id=user_id, role_id=role.id, created_at=datetime.now()))
    return True


def _reset_non_admin_users(db) -> Counter:
    _log("resetting non-admin users...")
    stats = Counter()
    admin = db.query(User).filter(User.username == "admin").first()
    admin_user_id = admin.user_id if admin else None

    query = db.query(User.user_id)
    if admin_user_id is not None:
        query = query.filter(User.user_id != admin_user_id)
    user_ids = [row[0] for row in query.all()]
    if not user_ids:
        _log("no non-admin users to reset")
        return stats
    _log(f"non-admin users to delete: {len(user_ids)}")

    policy_ids = [
        row[0]
        for row in db.query(DataPolicy.id)
        .filter(DataPolicy.subject_type == "USER", DataPolicy.subject_id.in_(user_ids))
        .all()
    ]
    if policy_ids:
        stats["data_policy_items_deleted"] = db.query(DataPolicyItem).filter(
            DataPolicyItem.policy_id.in_(policy_ids)
        ).delete(synchronize_session=False)
        stats["data_policies_deleted"] = db.query(DataPolicy).filter(
            DataPolicy.id.in_(policy_ids)
        ).delete(synchronize_session=False)

    stats["department_managers_cleared"] = db.query(Department).filter(
        Department.manager_user_id.in_(user_ids)
    ).update({Department.manager_user_id: None}, synchronize_session=False)
    stats["operation_logs_unlinked"] = db.query(OperationLog).filter(
        OperationLog.user_id.in_(user_ids)
    ).update({OperationLog.user_id: None}, synchronize_session=False)
    stats["system_logs_unlinked"] = db.query(SystemLog).filter(
        SystemLog.user_id.in_(user_ids)
    ).update({SystemLog.user_id: None}, synchronize_session=False)
    stats["login_logs_deleted"] = db.query(LoginLog).filter(
        LoginLog.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["user_department_posts_deleted"] = db.query(UserDepartmentPost).filter(
        UserDepartmentPost.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["user_roles_deleted"] = db.query(UserRole).filter(
        UserRole.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["user_identities_deleted"] = db.query(UserIdentity).filter(
        UserIdentity.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    stats["users_deleted"] = db.query(User).filter(
        User.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    return stats


def sync_wecom_contacts(
    *,
    apply: bool,
    from_file: str,
    reset_users: bool,
    reset_wecom_identities: bool,
    disable_missing_wecom_users: bool,
    grant_viewer_role: bool,
    default_password: str,
) -> int:
    corp_id = _get_env("WECOM_CORP_ID")
    if not corp_id:
        raise WeComSyncError("缺少 WECOM_CORP_ID")

    if from_file:
        department_ids: list[int] = []
        members = load_members_from_file(from_file)
    else:
        secret = _get_env("WECOM_CONTACT_SECRET") or _get_env("WECOM_APP_SECRET")
        if not secret:
            raise WeComSyncError("缺少 WECOM_CONTACT_SECRET 或 WECOM_APP_SECRET")
        _log("fetching Enterprise WeChat access token...")
        access_token = get_access_token(corp_id, secret)
        _log("access token fetched")
        department_ids = list_department_ids(access_token)
        members = list_members(access_token, department_ids)
    active_userids = {member.userid for member in members if _is_active_status(member.status)}

    db = SessionLocal()
    stats = Counter()
    try:
        stats["departments_seen"] = len(department_ids)
        stats["members_seen"] = len(members)
        stats["members_active"] = len(active_userids)

        if apply and reset_users:
            stats.update(_reset_non_admin_users(db))

        if apply and reset_wecom_identities:
            deleted = db.query(UserIdentity).filter(UserIdentity.identity_type == "wecom").delete(synchronize_session=False)
            stats["wecom_identities_deleted"] = deleted

        _log(f"syncing members into database: {len(members)}")
        for index, member in enumerate(members, start=1):
            if index == 1 or index % 20 == 0 or index == len(members):
                _log(f"sync progress: {index}/{len(members)}")
            if not _is_active_status(member.status):
                stats["members_inactive"] += 1

            if not apply:
                user = _find_user_for_member(db, member)
                stats["users_matched" if user else "users_would_create"] += 1
                continue

            user, created = _upsert_user(db, member, default_password)
            stats["users_created" if created else "users_updated"] += 1
            if _upsert_wecom_identity(db, user, corp_id, member):
                stats["wecom_identities_created"] += 1
            else:
                stats["wecom_identities_updated"] += 1
            if grant_viewer_role and _ensure_viewer_role(db, user.user_id):
                stats["viewer_roles_added"] += 1
            if _sync_user_department(db, user, member):
                stats["user_departments_created"] += 1

        if apply and disable_missing_wecom_users:
            identities = (
                db.query(UserIdentity)
                .filter(UserIdentity.identity_type == "wecom", UserIdentity.corp_id == corp_id)
                .all()
            )
            for identity in identities:
                if identity.wecom_user_id in active_userids:
                    continue
                user = db.query(User).filter(User.user_id == identity.user_id).first()
                if user and user.username != "admin":
                    user.status = "DISABLED"
                    user.is_active = False
                    user.updated_at = datetime.now()
                    stats["users_disabled_missing_wecom"] += 1

        if apply:
            db.commit()
        else:
            db.rollback()

        print("Enterprise WeChat contacts sync summary")
        print(f"mode: {'APPLY' if apply else 'DRY-RUN'}")
        print(f"corp_id: {corp_id}")
        print(f"from_file: {from_file or '-'}")
        print(f"reset_users: {reset_users}")
        print(f"reset_wecom_identities: {reset_wecom_identities}")
        print(f"disable_missing_wecom_users: {disable_missing_wecom_users}")
        print(f"grant_viewer_role: {grant_viewer_role}")
        print("stats: " + (", ".join(f"{key}={value}" for key, value in sorted(stats.items())) or "none"))
        if not apply:
            print("dry-run only; add --apply to write changes")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Enterprise WeChat contacts into ShopView users.")
    parser.add_argument("--from-file", default="", help="Import members from Enterprise WeChat exported .csv/.xlsx instead of API.")
    parser.add_argument("--apply", action="store_true", help="Write changes. Omit for dry-run.")
    parser.add_argument("--reset-users", action="store_true", help="Delete all non-admin users and related user auth/data policy rows before syncing.")
    parser.add_argument("--reset-wecom-identities", action="store_true", help="Delete existing wecom identities before syncing.")
    parser.add_argument("--disable-missing-wecom-users", action="store_true", help="Disable users bound to wecom ids no longer active.")
    parser.add_argument("--grant-viewer-role", action="store_true", help="Grant the contract viewer role to synced users.")
    parser.add_argument("--default-password", default=DEFAULT_ADMIN_PASSWORD, help="Initial password for newly-created users.")
    args = parser.parse_args()

    try:
        return sync_wecom_contacts(
            apply=args.apply,
            from_file=args.from_file,
            reset_users=args.reset_users,
            reset_wecom_identities=args.reset_wecom_identities,
            disable_missing_wecom_users=args.disable_missing_wecom_users,
            grant_viewer_role=args.grant_viewer_role,
            default_password=args.default_password,
        )
    except WeComSyncError as exc:
        print(f"同步失败: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
