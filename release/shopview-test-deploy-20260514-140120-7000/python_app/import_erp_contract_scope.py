#!/usr/bin/env python3
"""
Import ERP users and contract/business scope from Excel files.

Default mode is dry-run. Add --apply to write changes.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Run `pip install -r python_requirements.txt` first.", file=sys.stderr)
    raise

sys.path.append(str(Path(__file__).resolve().parent))

from models.database import SessionLocal
from models.models import CounterGroup, DataPolicy, DataPolicyItem, Department, Store, User, UserRole
from routers.auth import DEFAULT_ADMIN_PASSWORD, _hash_password
from routers.system_management import _ensure_contract_viewer_role


SOURCE_TYPE = "ERP"
SOURCE_SYSTEM = "erp_contract_scope"
BUSINESS_SCOPE_RESOURCE = "business_scope"
BUSINESS_SCOPE_ACTION = "view"


@dataclass
class ErpUser:
    username: str
    real_name: str
    range_text: str
    max_date: date | None
    raw: dict[str, str]


@dataclass
class ScopeBuildResult:
    scope_mode: str
    items: set[tuple[str, str]] = field(default_factory=set)
    skipped: list[tuple[str, str, str]] = field(default_factory=list)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_rows(path: str | Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = [_cell_text(value) for value in next(rows)]
    result = []
    for values in rows:
        row = {headers[index]: _cell_text(value) for index, value in enumerate(values) if index < len(headers)}
        if any(row.values()):
            result.append(row)
    workbook.close()
    return result


def _parse_date(value: str) -> date | None:
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value[:19], fmt).date()
        except ValueError:
            continue
    return None


def _split_ranges(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def load_erp_users(path: str | Path) -> list[ErpUser]:
    users = []
    for row in _load_rows(path):
        username = row.get("OAGH", "").strip()
        if not username:
            continue
        users.append(
            ErpUser(
                username=username,
                real_name=row.get("OANAME", "").strip(),
                range_text=row.get("OARANGE", "").strip() or "0",
                max_date=_parse_date(row.get("OAMAXDATE", "")),
                raw=row,
            )
        )
    return users


def load_scope_rows(path: str | Path) -> dict[str, list[dict[str, str]]]:
    by_role: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in _load_rows(path):
        role_id = row.get("OSRROLEID", "").strip()
        if role_id:
            by_role[role_id].append(row)
    return by_role


def _active_from_max_date(max_date: date | None) -> bool:
    return max_date is None or max_date >= date.today()


def build_scope(
    erp_user: ErpUser,
    scope_rows_by_role: dict[str, list[dict[str, str]]],
    store_by_code: dict[str, Store],
    dept_codes: set[str],
    group_codes: set[str],
) -> ScopeBuildResult:
    role_ids = _split_ranges(erp_user.range_text)
    if not role_ids or role_ids == ["0"] or "0" in role_ids:
        return ScopeBuildResult(scope_mode="ALL")

    result = ScopeBuildResult(scope_mode="CUSTOM")
    for role_id in role_ids:
        rows = scope_rows_by_role.get(role_id, [])
        if not rows:
            result.skipped.append((role_id, "", "missing_scope_role"))
            continue

        for row in rows:
            mfid = row.get("OSRMFID", "").strip()
            market = row.get("OSRMARKET", "").strip()
            if mfid == "0":
                store = store_by_code.get(market)
                if store:
                    result.items.add(("store", str(store.store_id)))
                else:
                    result.skipped.append((role_id, market, "missing_store_code"))
                continue

            if len(mfid) == 7:
                if mfid in dept_codes:
                    result.items.add(("department", mfid))
                else:
                    result.skipped.append((role_id, mfid, "missing_department_code"))
                continue

            if len(mfid) == 10:
                if mfid in group_codes:
                    result.items.add(("group", mfid))
                else:
                    result.skipped.append((role_id, mfid, "missing_group_code"))
                continue

            result.skipped.append((role_id, mfid or market, "unsupported_scope_code"))

    return result


def _delete_existing_erp_scope(db, user_id: int) -> int:
    policies = (
        db.query(DataPolicy)
        .filter(
            DataPolicy.subject_type == "USER",
            DataPolicy.subject_id == user_id,
            DataPolicy.resource_code == BUSINESS_SCOPE_RESOURCE,
            DataPolicy.action_code == BUSINESS_SCOPE_ACTION,
            DataPolicy.source_type == SOURCE_TYPE,
            DataPolicy.source_system == SOURCE_SYSTEM,
        )
        .all()
    )
    policy_ids = [policy.id for policy in policies]
    if policy_ids:
        db.query(DataPolicyItem).filter(DataPolicyItem.policy_id.in_(policy_ids)).delete(synchronize_session=False)
        db.query(DataPolicy).filter(DataPolicy.id.in_(policy_ids)).delete(synchronize_session=False)
    return len(policy_ids)


def _ensure_user(db, erp_user: ErpUser, default_password: str) -> tuple[User, bool]:
    user = db.query(User).filter(User.username == erp_user.username).first()
    is_active = _active_from_max_date(erp_user.max_date)
    if user:
        user.real_name = erp_user.real_name or user.real_name
        user.employee_no = erp_user.username
        user.status = "ACTIVE" if is_active else "DISABLED"
        user.is_active = is_active
        return user, False

    user = User(
        username=erp_user.username,
        password_hash=_hash_password(default_password),
        real_name=erp_user.real_name,
        employee_no=erp_user.username,
        role="user",
        status="ACTIVE" if is_active else "DISABLED",
        is_active=is_active,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(user)
    db.flush()
    return user, True


def _ensure_user_role(db, user_id: int, role_id: int) -> bool:
    exists = db.query(UserRole).filter(UserRole.user_id == user_id, UserRole.role_id == role_id).first()
    if exists:
        return False
    db.add(UserRole(user_id=user_id, role_id=role_id, created_at=datetime.now()))
    return True


def _insert_policy(db, user_id: int, scope: ScopeBuildResult, external_scope_id: str) -> None:
    policy = DataPolicy(
        subject_type="USER",
        subject_id=user_id,
        resource_code=BUSINESS_SCOPE_RESOURCE,
        action_code=BUSINESS_SCOPE_ACTION,
        scope_mode=scope.scope_mode,
        effect="ALLOW",
        priority=100,
        is_active=True,
        source_type=SOURCE_TYPE,
        source_system=SOURCE_SYSTEM,
        external_scope_id=external_scope_id,
        external_scope_name=f"ERP range {external_scope_id}",
        synced_at=datetime.now(),
    )
    db.add(policy)
    db.flush()
    for dimension_type, dimension_value in sorted(scope.items):
        db.add(
            DataPolicyItem(
                policy_id=policy.id,
                dimension_type=dimension_type,
                dimension_value=dimension_value,
                include_children=False,
            )
        )


def _format_counter(counter: Counter) -> str:
    return ", ".join(f"{key}={value}" for key, value in sorted(counter.items())) or "none"


def import_erp_contract_scope(users_xlsx: str, scope_xlsx: str, apply: bool, default_password: str) -> int:
    erp_users = load_erp_users(users_xlsx)
    scope_rows_by_role = load_scope_rows(scope_xlsx)

    db = SessionLocal()
    stats = Counter()
    skipped_reasons = Counter()
    skipped_samples: list[tuple[str, str, str, str]] = []

    try:
        store_by_code = {store.store_code: store for store in db.query(Store).all()}
        dept_codes = {row[0] for row in db.query(Department.dept_code).all()}
        group_codes = {row[0] for row in db.query(CounterGroup.group_code).all()}
        viewer_role = _ensure_contract_viewer_role(db)

        for erp_user in erp_users:
            scope = build_scope(erp_user, scope_rows_by_role, store_by_code, dept_codes, group_codes)
            stats["users_seen"] += 1
            stats[f"scope_{scope.scope_mode.lower()}"] += 1
            stats["scope_items"] += len(scope.items)
            for role_id, code, reason in scope.skipped:
                skipped_reasons[reason] += 1
                if len(skipped_samples) < 30:
                    skipped_samples.append((erp_user.username, role_id, code, reason))

            if not apply:
                continue

            user, created = _ensure_user(db, erp_user, default_password)
            stats["users_created" if created else "users_updated"] += 1
            if _ensure_user_role(db, user.user_id, viewer_role.id):
                stats["viewer_roles_added"] += 1
            stats["erp_policies_deleted"] += _delete_existing_erp_scope(db, user.user_id)

            if scope.scope_mode == "ALL" or scope.items:
                _insert_policy(db, user.user_id, scope, erp_user.range_text)
                stats["erp_policies_created"] += 1
            else:
                stats["users_without_policy"] += 1

        if apply:
            db.commit()
        else:
            db.rollback()

        print("ERP contract scope import summary")
        print(f"mode: {'APPLY' if apply else 'DRY-RUN'}")
        print(f"users file: {users_xlsx}")
        print(f"scope file: {scope_xlsx}")
        print(f"stats: {_format_counter(stats)}")
        print(f"skipped: {_format_counter(skipped_reasons)}")
        if skipped_samples:
            print("skipped samples:")
            for username, role_id, code, reason in skipped_samples:
                print(f"  user={username} role={role_id} code={code} reason={reason}")
        return 0
    except Exception as exc:
        db.rollback()
        print(f"Import failed: {exc}", file=sys.stderr)
        return 1
    finally:
        db.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import ERP users and contract scope from Excel.")
    parser.add_argument("--users-xlsx", required=True, help="Path to 用户表.xlsx")
    parser.add_argument("--scope-xlsx", required=True, help="Path to 数据范围标.xlsx")
    parser.add_argument("--default-password", default=DEFAULT_ADMIN_PASSWORD, help="Initial password for newly-created users")
    parser.add_argument("--apply", action="store_true", help="Write changes. Without this flag, only prints a dry-run summary.")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    raise SystemExit(import_erp_contract_scope(args.users_xlsx, args.scope_xlsx, args.apply, args.default_password))
