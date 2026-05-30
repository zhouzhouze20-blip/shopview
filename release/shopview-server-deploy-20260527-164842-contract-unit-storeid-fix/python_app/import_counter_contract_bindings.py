#!/usr/bin/env python3
"""
Import shop unit / counter group / contract bindings from the cabinet contract Excel.

Default mode is dry-run. Add --apply to write rows into business_unit_binding.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    import xlrd
except ImportError:
    print("Missing dependency: xlrd. Run `pip install -r python_requirements.txt` first.", file=sys.stderr)
    raise

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parent))

from models.database import SessionLocal


DEFAULT_SOURCE_PREFIX = "柜位合同1-5.xls"


@dataclass(frozen=True)
class BindingRow:
    sheet_name: str
    floor_text: str
    unit_code: str
    group_code: str
    group_name: str
    contract_id: str
    oa_contract_id: str
    area: Decimal | None
    unit_type: str
    drawing_area: Decimal | None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: Any) -> Decimal | None:
    text_value = _text(value)
    if not text_value:
        return None
    try:
        return Decimal(text_value)
    except Exception:
        return None


def _contract_id(value: Any) -> str:
    text_value = _text(value)
    if not text_value:
        return ""
    if text_value.isdigit() and len(text_value) < 8:
        return text_value.zfill(8)
    return text_value


def _normalize_code(value: str) -> str:
    return "".join(value.upper().split())


def _first_cell(header_index: dict[str, int], row: list[Any], *headers: str) -> Any:
    for header in headers:
        col_index = header_index.get(header)
        if col_index is not None and col_index < len(row):
            return row[col_index]
    return None


def _row_from_values(sheet_name: str, header_index: dict[str, int], row: list[Any]) -> BindingRow | None:
    unit_code = _text(_first_cell(header_index, row, "柜位编号"))
    group_code = _text(_first_cell(header_index, row, "柜组", "柜组编码"))
    contract_id = _contract_id(_first_cell(header_index, row, "富基合同编号"))
    if not unit_code and not group_code and not contract_id:
        return None
    return BindingRow(
        sheet_name=sheet_name,
        floor_text=_text(_first_cell(header_index, row, "楼层")),
        unit_code=unit_code,
        group_code=group_code,
        group_name=_text(_first_cell(header_index, row, "柜组名称")),
        contract_id=contract_id,
        oa_contract_id=_text(_first_cell(header_index, row, "OA合同编号")),
        area=_decimal(_first_cell(header_index, row, "面积(㎡)")),
        unit_type=_text(_first_cell(header_index, row, "柜位性质")),
        drawing_area=_decimal(_first_cell(header_index, row, "图纸面积")),
    )


def load_rows(path: str | Path) -> list[BindingRow]:
    path = Path(path)
    rows: list[BindingRow] = []
    if path.suffix.lower() == ".xlsx":
        if load_workbook is None:
            print("Missing dependency: openpyxl. Run `pip install -r python_requirements.txt` first.", file=sys.stderr)
            raise SystemExit(1)
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
            if len(values) < 2:
                continue
            headers = [_text(value) for value in values[0]]
            header_index = {header: index for index, header in enumerate(headers) if header}
            for row in values[1:]:
                binding_row = _row_from_values(sheet.title, header_index, row)
                if binding_row is not None:
                    rows.append(binding_row)
        return rows

    workbook = xlrd.open_workbook(str(path))
    for sheet in workbook.sheets():
        if sheet.nrows < 2:
            continue
        headers = [_text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        header_index = {header: index for index, header in enumerate(headers) if header}
        for row_index in range(1, sheet.nrows):
            row = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
            binding_row = _row_from_values(sheet.name, header_index, row)
            if binding_row is not None:
                rows.append(binding_row)
    return rows


def _binding_remark(row: BindingRow, source_prefix: str = DEFAULT_SOURCE_PREFIX) -> str:
    parts = [
        source_prefix,
        f"sheet={row.sheet_name}",
        f"floor={row.floor_text}",
        f"unit={row.unit_code}",
    ]
    if row.group_code:
        parts.append(f"group={row.group_code}")
    if row.group_name:
        parts.append(f"group_name={row.group_name}")
    if row.oa_contract_id:
        parts.append(f"oa={row.oa_contract_id}")
    if row.unit_type:
        parts.append(f"type={row.unit_type}")
    if row.area is not None:
        parts.append(f"area={row.area}")
    if row.drawing_area is not None:
        parts.append(f"drawing_area={row.drawing_area}")
    return "; ".join(parts)


def import_bindings(
    path: str | Path,
    apply: bool = False,
    replace_source: bool = False,
    skip_existing: bool = False,
) -> int:
    path = Path(path)
    source_prefix = path.name or DEFAULT_SOURCE_PREFIX
    source_rows = load_rows(path)
    stats = Counter()
    seen_keys: set[tuple[int, str]] = set()

    db = SessionLocal()
    try:
        unit_rows = db.execute(text("SELECT id, unit_code FROM business_units")).mappings().all()
        unit_by_code = {_normalize_code(row["unit_code"]): int(row["id"]) for row in unit_rows}

        group_rows = db.execute(text("SELECT group_id, group_code FROM counter_groups")).mappings().all()
        group_by_code = {_normalize_code(row["group_code"]): int(row["group_id"]) for row in group_rows}

        if apply and replace_source:
            deleted = db.execute(
                text("DELETE FROM business_unit_binding WHERE remark LIKE :source_prefix"),
                {"source_prefix": f"{source_prefix}%"},
            ).rowcount
            stats["deleted_source_rows"] = deleted or 0

        for row in source_rows:
            stats["excel_rows"] += 1
            if not row.unit_code:
                stats["missing_unit_code"] += 1
                continue
            if not row.contract_id:
                stats["missing_contract_id"] += 1
                continue

            shop_unit_id = unit_by_code.get(_normalize_code(row.unit_code))
            if not shop_unit_id:
                stats["unmatched_shop_unit"] += 1
                continue

            dedupe_key = (shop_unit_id, row.contract_id)
            if dedupe_key in seen_keys:
                stats["duplicate_excel_binding"] += 1
                continue
            seen_keys.add(dedupe_key)

            counter_group_id = group_by_code.get(_normalize_code(row.group_code)) if row.group_code else None
            if row.group_code and counter_group_id is None:
                stats["unmatched_counter_group"] += 1

            existing = db.execute(
                text(
                    """
                    SELECT id
                    FROM business_unit_binding
                    WHERE shop_unit_id = :shop_unit_id
                      AND upper(trim(contract_id)) = upper(trim(:contract_id))
                    LIMIT 1
                    """
                ),
                {"shop_unit_id": shop_unit_id, "contract_id": row.contract_id},
            ).fetchone()

            values = {
                "shop_unit_id": shop_unit_id,
                "counter_group_id": counter_group_id,
                "contract_id": row.contract_id,
                "brand_id": row.group_name or None,
                "business_type": row.unit_type or None,
                "remark": _binding_remark(row, source_prefix),
            }
            if existing:
                if skip_existing:
                    stats["skipped_existing"] += 1
                    continue
                stats["would_update" if not apply else "updated"] += 1
                if apply:
                    db.execute(
                        text(
                            """
                            UPDATE business_unit_binding
                            SET counter_group_id = :counter_group_id,
                                brand_id = :brand_id,
                                business_type = :business_type,
                                status = 'ACTIVE',
                                remark = :remark,
                                updated_at = NOW()
                            WHERE id = :id
                            """
                        ),
                        {**values, "id": existing.id},
                    )
                continue

            stats["would_insert" if not apply else "inserted"] += 1
            if apply:
                db.execute(
                    text(
                        """
                        INSERT INTO business_unit_binding (
                            shop_unit_id,
                            counter_group_id,
                            brand_id,
                            contract_id,
                            business_type,
                            is_primary,
                            status,
                            remark,
                            created_at,
                            updated_at
                        )
                        VALUES (
                            :shop_unit_id,
                            :counter_group_id,
                            :brand_id,
                            :contract_id,
                            :business_type,
                            TRUE,
                            'ACTIVE',
                            :remark,
                            NOW(),
                            NOW()
                        )
                        """
                    ),
                    values,
                )

        if apply:
            db.commit()
        else:
            db.rollback()

        print("Import summary")
        for key, value in sorted(stats.items()):
            print(f"{key}: {value}")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Import shop unit / counter group / contract bindings from Excel.")
    parser.add_argument("excel_path", help="Path to .xls workbook")
    parser.add_argument("--apply", action="store_true", help="Write changes. Default is dry-run.")
    parser.add_argument(
        "--replace-source",
        action="store_true",
        help="Delete previous rows whose remark starts with the source filename before importing.",
    )
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="Skip existing shop unit + contract bindings instead of updating them.",
    )
    args = parser.parse_args()
    return import_bindings(
        args.excel_path,
        apply=args.apply,
        replace_source=args.replace_source,
        skip_existing=args.skip_existing,
    )


if __name__ == "__main__":
    raise SystemExit(main())
